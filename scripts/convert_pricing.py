"""
Convert the FOB-CIF pricing Excel to a JS data module.
Usage: python scripts/convert_pricing.py
Output: src/data/countryPricingData.js
"""
import openpyxl
import json
import re
import os
import sys

EXCEL_PATH = r"c:\Users\28307\Desktop\拉美国家(矿机)-FOB-CIF 价格汇总表(补充信息)-2025.10.xlsx"
OUTPUT_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "src", "data", "countryPricingData.js")

# Category mapping based on product name keywords
CATEGORY_MAP = [
    ("电动轮矿卡", "矿卡"),
    ("纯电矿卡", "矿卡"),
    ("宽体车", "宽体车"),
    ("矿挖", "矿挖"),
    ("电挖", "矿挖"),
    ("鄂破", "破碎设备"),
    ("反击破", "破碎设备"),
    ("圆锥破", "破碎设备"),
    ("筛分站", "筛分设备"),
    ("重型筛分站", "筛分设备"),
    ("潜孔钻机", "钻机"),
]

# Keywords for identifying key columns
MODEL_KEYS = ["机型"]
NAME_KEYS = ["名称"]
DIM_KEYS = ["外形尺寸"]
WEIGHT_KEYS = ["运输重量"]
FOB_KEYS = ["FOB"]
FREIGHT_KEYS = ["海运费"]
INSURANCE_KEYS = ["保险"]
CIF_KEYS = ["CIF"]
DDP_KEYS = ["DDP价", "DDP"]
EXW_KEYS = ["EXW"]
VAT_KEYS = ["增值税", "销售增值"]


def clean_header(h):
    """Normalize header text by removing newlines and extra spaces."""
    return re.sub(r'\s+', ' ', str(h).strip())


def find_column(headers, keywords):
    """Find the first column index whose header contains any of the keywords."""
    for kw in keywords:
        for i, h in enumerate(headers):
            if kw in h:
                return i
    return -1


def derive_category(name):
    """Map product name to a broad category."""
    for kw, cat in CATEGORY_MAP:
        if kw in (name or ""):
            return cat
    return "其他"


def is_product_row(row, model_col, name_col):
    """Check if a row is a product row (has model or name)."""
    if model_col >= 0 and model_col < len(row):
        v = str(row[model_col]).strip()
        if v and not v.startswith("说明") and not re.match(r'^\d+\.', v):
            return True
    if name_col >= 0 and name_col < len(row):
        v = str(row[name_col]).strip()
        if v and v not in ("", "名称") and not re.match(r'^\d+\.', v):
            return True
    return False


def parse_directory_sheet(ws):
    """Parse the directory sheet into a list of country info dicts."""
    countries = []
    # Header row is at index 2 (0-based), data starts at row 3
    # Columns: col1=region, col2=seq, col3=country, col4=person, col5=pricingModel, col10=notes
    for i, row in enumerate(ws.iter_rows(min_row=4, max_row=35, values_only=True)):
        vals = [str(c).strip() if c is not None else "" for c in row]
        col1, col2, col3, col4, col5, col10 = "", "", "", "", "", ""
        if len(vals) > 1: col1 = vals[1]
        if len(vals) > 2: col2 = vals[2]
        if len(vals) > 3: col3 = vals[3]
        if len(vals) > 4: col4 = vals[4]
        if len(vals) > 5: col5 = vals[5]
        if len(vals) > 10: col10 = vals[10]

        if not col2.isdigit() or not col3:
            break
        countries.append({
            "region": col1,
            "seq": int(col2),
            "country": col3.replace("\xa0", "").replace("　", ""),
            "person": col4,
            "pricingModel": col5,
            "notes": col10,
        })
    return countries


def parse_country_sheet(ws):
    """Parse a country product sheet. Returns (headers, products)."""
    raw_rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        raw_rows.append([str(c).strip() if c is not None else "" for c in row])

    # Find header row: first row containing "机型" or "序号" and "名称"
    header_idx = -1
    for i, row in enumerate(raw_rows):
        text = " ".join(row)
        if "机型" in text or ("序号" in row and "名称" in text):
            header_idx = i
            break

    if header_idx < 0:
        header_idx = 0

    headers = [clean_header(h) for h in raw_rows[header_idx]]

    # Find key column indices
    model_col = find_column(headers, MODEL_KEYS)
    name_col = find_column(headers, NAME_KEYS)
    dim_col = find_column(headers, DIM_KEYS)
    weight_col = find_column(headers, WEIGHT_KEYS)
    fob_col = find_column(headers, FOB_KEYS)
    freight_col = find_column(headers, FREIGHT_KEYS)
    insurance_col = find_column(headers, INSURANCE_KEYS)
    cif_col = find_column(headers, CIF_KEYS)
    ddp_col = find_column(headers, DDP_KEYS)
    exw_col = find_column(headers, EXW_KEYS)
    vat_col = find_column(headers, VAT_KEYS)

    products = []
    seq = 0
    for row in raw_rows[header_idx + 1:]:
        # Stop at notes/empty rows
        text = " ".join(row[:3])
        if not text.strip():
            break
        if text.startswith("说明") or re.match(r'^\d+\.', text):
            break

        if not is_product_row(row, model_col, name_col):
            continue

        seq += 1
        model = row[model_col] if model_col >= 0 and model_col < len(row) else ""
        name = row[name_col] if name_col >= 0 and name_col < len(row) else ""
        category = derive_category(name)

        # Build all-fields map
        all_fields = {}
        for j, h in enumerate(headers):
            if j < len(row) and row[j]:
                all_fields[h] = row[j]

        product = {
            "seq": seq,
            "model": model,
            "name": name,
            "category": category,
            "dimensions": row[dim_col] if dim_col >= 0 and dim_col < len(row) else "",
            "weight": row[weight_col] if weight_col >= 0 and weight_col < len(row) else "",
            "fob": row[fob_col] if fob_col >= 0 and fob_col < len(row) else "",
            "oceanFreight": row[freight_col] if freight_col >= 0 and freight_col < len(row) else "",
            "insurance": row[insurance_col] if insurance_col >= 0 and insurance_col < len(row) else "",
            "cif": row[cif_col] if cif_col >= 0 and cif_col < len(row) else "",
            "ddp": row[ddp_col] if ddp_col >= 0 and ddp_col < len(row) else "",
            "exw": row[exw_col] if exw_col >= 0 and exw_col < len(row) else "",
            "vat": row[vat_col] if vat_col >= 0 and vat_col < len(row) else "",
            "allFields": all_fields,
        }
        products.append(product)

    return headers, products


def main():
    print(f"Reading: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # First sheet is always the directory sheet
    dir_sheet = wb.sheetnames[0]
    print(f"Directory sheet: {dir_sheet}")
    countries_info = parse_directory_sheet(wb[dir_sheet])
    print(f"Found {len(countries_info)} countries in directory")

    # Build lookup for country info
    country_info_map = {c["country"]: c for c in countries_info}

    # Parse each country sheet
    result = []
    for name in wb.sheetnames:
        if name == dir_sheet:
            continue
        print(f"Parsing: {name}")
        headers, products = parse_country_sheet(wb[name])
        info = country_info_map.get(name, {})
        result.append({
            "country": name,
            "region": info.get("region", ""),
            "person": info.get("person", ""),
            "pricingModel": info.get("pricingModel", ""),
            "notes": info.get("notes", ""),
            "headers": headers,
            "products": products,
        })
        print(f"  -> {len(products)} products, {len(headers)} columns")

    # Write JS module
    json_str = json.dumps(result, ensure_ascii=False, indent=2)
    js_content = f"""// Auto-generated by scripts/convert_pricing.py
// Source: 拉美国家(矿机)-FOB-CIF 价格汇总表(补充信息)-2025.10.xlsx
// Generated: {__import__('datetime').datetime.now().isoformat()}
const countryPricingData = {json_str};
export default countryPricingData;
"""

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write(js_content)

    print(f"\nWritten: {OUTPUT_PATH}")
    total_products = sum(len(c["products"]) for c in result)
    print(f"Total: {len(result)} countries, {total_products} products")


if __name__ == "__main__":
    main()
